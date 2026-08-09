#!/usr/bin/env python3
"""
=============================================================================
GENERADOR MASIVO DE DICCIONARIOS DE DATOS DESDE SHAPEFILES
=============================================================================
Escanea recursivamente una carpeta raíz, detecta todos los shapefiles (.shp)
y genera, por cada uno, un Excel con el diccionario de datos (estructura de
campos, dominios de valores, tipo de dato y longitud real).

Puntos clave:
    - La longitud de los campos de tipo texto se lee directamente del
      esquema nativo del shapefile (.dbf) usando `pyshp`, en vez de usar
      un diccionario fijo o un valor por defecto (254). Esto garantiza que
      el valor mostrado en el Excel coincida con el que se ve en QGIS
      (Propiedades de la capa > Campos > Longitud).
    - El título del Excel usa el valor del campo DESCR_PLAN. Si ese valor
      comienza con "PLAN ESPECIFICO" (con o sin tilde), se abrevia a "PE"
      conservando el resto del texto.

Uso:
    python generar_diccionarios.py /ruta/a/la/carpeta/raiz
    python generar_diccionarios.py /ruta/a/la/carpeta/raiz --salida /ruta/salida
=============================================================================
"""

import argparse
import os
import sys
import traceback

import shapefile  # pyshp -> pip install pyshp
import geopandas as gpd
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# =============================================================================
# 1. CONFIGURACIÓN GENERAL
# =============================================================================

# Fuente tipográfica usada en el Excel. Si no está instalada en el sistema,
# Excel la sustituirá automáticamente por una fuente similar al abrir el
# archivo; esto no afecta la generación del reporte.
FUENTE_EXCEL = "Swis721 Cn BT"

# Respaldo: solo se usa si un campo de texto no aparece en el esquema del
# shapefile (caso excepcional). La fuente de verdad es el .dbf vía pyshp.
LONGITUDES_RESPALDO = {
    "REFOREST": 50,
    "PLAN": 15,
    "DESCR_PLAN": 30,
    "FASE_PLAN": 10,
    "FUENTE_PRI": 30,
}

MAPA_TIPOS = {
    "object": "Text", "string": "Text", "str": "Text",
    "int64": "Entero", "int32": "Entero",
    "float64": "Double", "float32": "Double",
}

MAPA_GEOMETRIAS = {
    "Point": "PUNTO", "MultiPoint": "PUNTO",
    "LineString": "POLILINEA", "MultiLineString": "POLILINEA",
    "Polygon": "POLIGONO", "MultiPolygon": "POLIGONO",
}

HEADERS = ["NOMBRE DE CAMPO", "DOMINIO O VALORES POSIBLES (TIPOLOGÍA)",
           "DESCRIPCIÓN", "TIPO DE DATO", "EXTENSION"]


# =============================================================================
# 2. ESTILOS DE EXCEL (openpyxl)
# =============================================================================

class Estilos:
    std = Font(name=FUENTE_EXCEL, size=11)
    bold = Font(name=FUENTE_EXCEL, size=11, bold=True)
    italic = Font(name=FUENTE_EXCEL, size=11, italic=True)
    white = Font(name=FUENTE_EXCEL, size=11, bold=True, color="FFFFFF")
    titulo = Font(name=FUENTE_EXCEL, size=14, bold=True)

    fill_verde = PatternFill(start_color="548135", fill_type="solid")
    fill_verde_claro = PatternFill(start_color="E2EFD9", fill_type="solid")

    borde = Border(*[Side(border_style="thin", color="000000")] * 4)
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)


# =============================================================================
# 3. FUNCIONES DE APOYO
# =============================================================================

def leer_esquema_dbf(ruta_sh):
    """
    Lee el esquema nativo del .dbf usando pyshp (sin depender de GDAL/fiona).
    Devuelve un diccionario {nombre_campo: ancho_caracteres} solo para
    campos de tipo texto ('C' = Character en la especificación dBASE).
    """
    lector = shapefile.Reader(ruta_sh)
    # El primer elemento de sf.fields es siempre ('DeletionFlag', ...) y se descarta
    campos = lector.fields[1:]
    return {
        nombre: ancho
        for nombre, tipo, ancho, _decimales in campos
        if tipo == "C"
    }


def obtener_ancho_real_texto(esquema_anchos_texto, nombre_campo):
    """
    Extrae el ancho real de un campo de texto según el esquema nativo del
    shapefile (.dbf), leído con pyshp.
    Devuelve None si el campo no es de tipo texto o no está en el esquema.
    """
    ancho = esquema_anchos_texto.get(nombre_campo)
    return str(ancho) if ancho else None


def calcular_extension(tipo_dato, columna, esquema_anchos_texto):
    """Determina el valor de la columna EXTENSION según el tipo de dato."""
    if tipo_dato == "Text":
        ancho_real = obtener_ancho_real_texto(esquema_anchos_texto, columna)
        if ancho_real:
            return ancho_real
        # Respaldo si el campo no aparece en el esquema del shapefile
        return str(LONGITUDES_RESPALDO.get(columna, 254))
    elif tipo_dato == "Double":
        return "-"
    else:
        return ""


def obtener_tipo_geometria(gdf):
    """Obtiene el tipo de geometría dominante del GeoDataFrame, en español."""
    geometrias_validas = gdf.geometry.geom_type.dropna()
    if geometrias_validas.empty:
        return ""
    return MAPA_GEOMETRIAS.get(str(geometrias_validas.iloc[0]), "")


def obtener_fuente_primaria(gdf):
    """Extrae los valores únicos del campo FUENTE_PRI, si existe."""
    if "FUENTE_PRI" not in gdf.columns:
        return "NINGUNA"
    fuentes = sorted([
        str(v) for v in gdf["FUENTE_PRI"].dropna().unique().tolist()
        if str(v).strip() != ""
    ])
    return ", ".join(fuentes) if fuentes else "NINGUNA"


def abreviar_descr_plan(valor):
    """
    Si el valor de DESCR_PLAN comienza con 'PLAN ESPECIFICO' (con o sin
    tilde en la Í), lo abrevia a 'PE' conservando el resto del texto.
    Ejemplo: 'PLAN ESPECIFICO DE RENOVACION URBANA' -> 'PE DE RENOVACION URBANA'
    """
    valor = valor.strip()
    prefijos = ("PLAN ESPECIFICO", "PLAN ESPECÍFICO")
    valor_mayus = valor.upper()
    for prefijo in prefijos:
        if valor_mayus.startswith(prefijo):
            resto = valor[len(prefijo):]
            return f"PE{resto}"
    return valor


def obtener_titulo_desde_descr_plan(gdf, valor_por_defecto):
    """
    Obtiene el texto a usar en el título del Excel a partir del campo
    DESCR_PLAN. Si el campo no existe o no tiene valores, se usa el
    nombre limpio del shapefile como respaldo.
    """
    if "DESCR_PLAN" not in gdf.columns:
        return valor_por_defecto

    valores = sorted(set(
        str(v).strip() for v in gdf["DESCR_PLAN"].dropna().unique().tolist()
        if str(v).strip() != ""
    ))
    if not valores:
        return valor_por_defecto

    # Se asume un único valor de DESCR_PLAN por capa; se toma el primero
    return abreviar_descr_plan(valores[0])


def construir_estructura_campos(gdf, esquema_anchos_texto):
    """
    Analiza cada columna del GeoDataFrame (excepto 'geometry') y arma la
    fila de datos: [nombre, dominio, descripción, tipo, extension, n_valores]
    """
    estructura = []
    for columna in [c for c in gdf.columns if c != "geometry"]:
        tipo_dato = MAPA_TIPOS.get(str(gdf[columna].dtype), str(gdf[columna].dtype))
        valores_unicos = sorted([str(v) for v in gdf[columna].dropna().unique().tolist()])
        n_valores = len(valores_unicos)

        extension = calcular_extension(tipo_dato, columna, esquema_anchos_texto)

        # Construcción del texto de dominio de valores
        if tipo_dato in ("Double", "Entero") or (tipo_dato == "Text" and n_valores > 15):
            dominio = f"Ejemplo: {valores_unicos[0]}" if valores_unicos else "Ejemplo: "
        elif tipo_dato == "Text":
            dominio = "\n".join([f"({i:02d}) {v}" for i, v in enumerate(valores_unicos, start=1)])
        else:
            if n_valores > 15:
                dominio = f"{', '.join(valores_unicos[:15])}... ({n_valores} valores)"
            else:
                dominio = ", ".join(valores_unicos)

        estructura.append([columna, dominio, "", tipo_dato, extension, n_valores])

    return estructura


def sanear_nombre_hoja(nombre):
    """
    Ajusta un texto para que sea válido como nombre de hoja de Excel:
    - Máximo 31 caracteres.
    - No permite los caracteres: : \\ / ? * [ ]
    """
    caracteres_invalidos = r'[]:\\/?*'
    nombre_limpio = "".join(c for c in nombre if c not in caracteres_invalidos)
    nombre_limpio = nombre_limpio.strip()
    return nombre_limpio[:31] if nombre_limpio else "Estructura"


def escribir_excel(ruta_xl, titulo_texto, nombre_limpio, f_class_sin_ext, geom_val, f_pri, estructura):
    """Genera el archivo Excel con el diseño corporativo del diccionario de datos."""
    metadatos = [
        ["Objeto:", nombre_limpio],
        ["Geometría:", geom_val],
        ["Grupo de objetos:", nombre_limpio],
        ["Feature class:", f_class_sin_ext],
        ["Tabla:", f"TB_{f_class_sin_ext}"],
        ["Fuente primaria:", f_pri],
        ["Fuente secundaria:", "NINGUNA"],
        ["Uso:", "Planes para el Acondicionamiento Territorial y Desarrollo Urbano (D.S. N° 012-2022-VIVIENDA)"],
    ]

    with pd.ExcelWriter(ruta_xl, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title=sanear_nombre_hoja(f_class_sin_ext))
        if "Sheet" in writer.book.sheetnames:
            writer.book.remove(writer.book["Sheet"])

        # --- Título principal (fila 1) ---
        ws.merge_cells("B1:F1")
        ws["B1"] = f"DICCIONARIO DE DATOS - {titulo_texto}"
        ws["B1"].font = Estilos.titulo
        ws["B1"].alignment = Estilos.align_centro

        # --- Metadatos (filas 3 a 10) ---
        for i, (clave, valor) in enumerate(metadatos, start=3):
            ws.cell(row=i, column=2, value=clave).font = Estilos.bold
            ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)
            ws.cell(row=i, column=3, value=valor)
            for c in range(2, 7):
                celda = ws.cell(row=i, column=c)
                celda.border = Estilos.borde
                celda.font = Estilos.bold
                celda.alignment = Estilos.align_izq

        # --- Encabezados de tabla (fila 11) ---
        for c_idx, encabezado in enumerate(HEADERS, start=2):
            celda = ws.cell(row=11, column=c_idx, value=encabezado)
            celda.font = Estilos.white
            celda.fill = Estilos.fill_verde
            celda.border = Estilos.borde
            celda.alignment = Estilos.align_centro

        # --- Cuerpo de la tabla (fila 12 en adelante) ---
        for r_idx, fila in enumerate(estructura, start=12):
            tipo_dato, n_valores = fila[3], fila[5]
            es_numerico = tipo_dato in ("Double", "Entero")
            es_texto_corto = (tipo_dato == "Text" and n_valores <= 15)
            es_texto_largo = (tipo_dato == "Text" and n_valores > 15)

            for c_idx, valor in enumerate(fila[:5], start=2):
                celda = ws.cell(row=r_idx, column=c_idx, value=valor)
                celda.border = Estilos.borde

                # Fuente según columna y condición
                if c_idx == 2 or (c_idx == 3 and es_texto_corto):
                    celda.font = Estilos.bold
                elif c_idx == 3 and (es_numerico or es_texto_largo):
                    celda.font = Estilos.italic
                else:
                    celda.font = Estilos.std

                # Alineación según columna y condición
                if c_idx in (2, 4) or (c_idx == 3 and (es_numerico or es_texto_largo)):
                    celda.alignment = Estilos.align_izq
                else:
                    celda.alignment = Estilos.align_centro

                # Relleno de columnas fijas del cuerpo
                if c_idx in (2, 5, 6) or (c_idx == 3 and not (es_numerico or es_texto_largo)):
                    celda.fill = Estilos.fill_verde_claro

        # --- Dimensionamiento de columnas ---
        anchos_columna = {"B": 22.0, "C": 45.0, "D": 35.0, "E": 16.0, "F": 14.0}
        for letra, ancho in anchos_columna.items():
            ws.column_dimensions[letra].width = ancho


# =============================================================================
# 4. PROCESAMIENTO DE UN SHAPEFILE INDIVIDUAL
# =============================================================================

def procesar_shapefile(ruta_sh, raiz, carpeta_salida=None):
    """Procesa un único shapefile y genera su Excel de diccionario de datos.

    Por defecto el Excel se guarda en la carpeta padre de `raiz` (mismo
    comportamiento que la versión original). Si se indica `carpeta_salida`,
    todos los Excel se guardan ahí en su lugar.
    """
    f_class_sin_ext = os.path.splitext(os.path.basename(ruta_sh))[0]
    destino = carpeta_salida if carpeta_salida else os.path.dirname(raiz)
    os.makedirs(destino, exist_ok=True)
    nombre_excel = f"DD_{f_class_sin_ext}.xlsx"
    ruta_xl = os.path.join(destino, nombre_excel)

    print(f"Shapefile encontrado: {ruta_sh}")
    print(f"Generando Excel -> {ruta_xl}\n")

    # Lectura de datos con geopandas (valores, tipos generales)
    gdf = gpd.read_file(ruta_sh)

    # Lectura del esquema nativo del .dbf con pyshp (longitudes reales de texto)
    esquema_anchos_texto = leer_esquema_dbf(ruta_sh)

    nombre_limpio = f_class_sin_ext.replace("_", " ")
    geom_val = obtener_tipo_geometria(gdf)
    f_pri = obtener_fuente_primaria(gdf)
    estructura = construir_estructura_campos(gdf, esquema_anchos_texto)

    # El título usa el valor de DESCR_PLAN (abreviando "PLAN ESPECIFICO" a "PE");
    # si el campo no existe o está vacío, se usa el nombre limpio como respaldo.
    titulo_texto = obtener_titulo_desde_descr_plan(gdf, valor_por_defecto=nombre_limpio)

    escribir_excel(ruta_xl, titulo_texto, nombre_limpio, f_class_sin_ext, geom_val, f_pri, estructura)
    print("✔ ¡Éxito! Archivo guardado correctamente.\n")


# =============================================================================
# 5. BUCLE MAESTRO (RECORRIDO RECURSIVO DE CARPETAS)
# =============================================================================

def generar_diccionarios(carpeta_raiz, carpeta_salida=None):
    """Recorre `carpeta_raiz` de forma recursiva y genera un Excel por cada .shp."""
    print("Iniciando el escaneo masivo y generación de diccionarios de datos...\n")

    encontrados = 0
    errores = 0

    for raiz, _directorios, archivos in os.walk(carpeta_raiz):
        for archivo in archivos:
            if archivo.lower().endswith(".shp"):
                encontrados += 1
                ruta_sh = os.path.join(raiz, archivo)
                try:
                    procesar_shapefile(ruta_sh, raiz, carpeta_salida)
                except Exception as e:
                    errores += 1
                    print(f"❌ Error al procesar el archivo {archivo}: {type(e).__name__}: {e}")
                    print(traceback.format_exc())
                    print()

    print(f"Proceso masivo completado. {encontrados - errores}/{encontrados} diccionarios generados correctamente.")
    return errores == 0


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Genera diccionarios de datos en Excel a partir de shapefiles, "
                    "recorriendo recursivamente una carpeta raíz."
    )
    parser.add_argument(
        "carpeta_raiz",
        help="Carpeta raíz donde buscar shapefiles (.shp) de forma recursiva.",
    )
    parser.add_argument(
        "--salida", "-o",
        dest="carpeta_salida",
        default=None,
        help="Carpeta donde guardar los Excel generados. "
             "Por defecto se guardan en la carpeta padre de cada shapefile.",
    )
    return parser.parse_args(argv)


def main():
    args = parse_args()

    if not os.path.isdir(args.carpeta_raiz):
        print(f"❌ La carpeta raíz no existe: {args.carpeta_raiz}")
        sys.exit(1)

    exito = generar_diccionarios(args.carpeta_raiz, args.carpeta_salida)
    sys.exit(0 if exito else 1)


if __name__ == "__main__":
    main()
